"""Export services: PDF reports and CSV exports."""
import csv
import io
import logging
from datetime import datetime
from typing import List, Dict, Optional

logger = logging.getLogger(__name__)


def generate_pdf_report(
    student_name: str,
    student_email: str,
    exam_title: str,
    subject_name: str,
    questions_and_scores: List[Dict],
    total_score: float,
    total_marks: float,
) -> bytes:
    """Generate a student report card as PDF.
    
    Args:
        questions_and_scores: List of dicts with keys:
            question_text, model_answer, student_answer, score, marks, feedback
    
    Returns:
        PDF bytes.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=22, spaceAfter=6)
    subtitle_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=12, alignment=TA_CENTER, textColor=colors.grey)
    heading_style = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=14, spaceAfter=4)
    body_style = ParagraphStyle("Body2", parent=styles["Normal"], fontSize=10, leading=14)
    feedback_style = ParagraphStyle("FB", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#4F46E5"), leftIndent=10)

    elements = []
    percentage = round((total_score / total_marks) * 100, 2) if total_marks > 0 else 0

    # Header
    elements.append(Paragraph("Student Report Card", title_style))
    elements.append(Paragraph("Smart Exam Answer Checker", subtitle_style))
    elements.append(Spacer(1, 8*mm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#4F46E5")))
    elements.append(Spacer(1, 4*mm))

    # Student info table
    info = [
        ["Student:", student_name],
        ["Email:", student_email],
        ["Exam:", exam_title],
        ["Subject:", subject_name],
        ["Date:", datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]
    t = Table(info, colWidths=[30*mm, 120*mm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 6*mm))

    # Score summary
    grade_color = "#10B981" if percentage >= 75 else ("#F59E0B" if percentage >= 50 else "#EF4444")
    summary = [
        ["Total Score", f"{total_score} / {total_marks}"],
        ["Percentage", f"{percentage}%"],
    ]
    st = Table(summary, colWidths=[50*mm, 50*mm])
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#d1d5db")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 8*mm))

    # Question-by-question breakdown
    elements.append(Paragraph("Detailed Results", heading_style))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    elements.append(Spacer(1, 3*mm))

    for i, q in enumerate(questions_and_scores, 1):
        elements.append(Paragraph(f"<b>Question {i}:</b> {q.get('question_text', '')[:200]}", body_style))
        elements.append(Paragraph(f"<b>Your Answer:</b> {q.get('student_answer', '')[:300]}", body_style))
        score_val = q.get('score', 0)
        marks_val = q.get('marks', 0)
        elements.append(Paragraph(f"<b>Score:</b> {score_val} / {marks_val}", body_style))
        fb = q.get('feedback', '')
        if fb:
            elements.append(Paragraph(f"Feedback: {fb}", feedback_style))
        elements.append(Spacer(1, 4*mm))

    doc.build(elements)
    return buf.getvalue()


def generate_csv_export(rows: List[Dict], headers: Optional[List[str]] = None) -> str:
    """Generate CSV content for teacher marks export.
    
    Args:
        rows: List of dicts with student data.
        headers: Optional custom column headers. If None, uses dict keys.
    
    Returns:
        CSV string.
    """
    if not rows:
        return ""

    if headers is None:
        headers = list(rows[0].keys())

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue()


def generate_exam_pdf(
    exam_title: str,
    subject_name: str,
    questions: List[Dict],
    total_marks: float,
    time_limit_minutes: Optional[int] = None,
) -> bytes:
    """Generate a printable exam paper as PDF.
    
    Args:
        questions: List of dicts with keys: question_text, marks
    
    Returns:
        PDF bytes.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("ETitle", parent=styles["Title"], fontSize=20, spaceAfter=4)
    subtitle_style = ParagraphStyle("ESub", parent=styles["Normal"], fontSize=11, alignment=TA_CENTER, textColor=colors.grey)
    q_style = ParagraphStyle("Q", parent=styles["Normal"], fontSize=11, leading=15, spaceAfter=2)
    marks_style = ParagraphStyle("M", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#4F46E5"))
    answer_style = ParagraphStyle("A", parent=styles["Normal"], fontSize=10, leftIndent=10, textColor=colors.grey)

    elements = []
    elements.append(Paragraph(exam_title, title_style))
    elements.append(Paragraph(f"{subject_name} | Total: {total_marks} marks" +
                              (f" | Time: {time_limit_minutes} min" if time_limit_minutes else ""), subtitle_style))
    elements.append(Spacer(1, 4*mm))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#4F46E5")))
    elements.append(Spacer(1, 6*mm))

    elements.append(Paragraph("<b>Instructions:</b> Answer all questions. Write clearly.", q_style))
    elements.append(Spacer(1, 6*mm))

    for i, q in enumerate(questions, 1):
        marks = q.get("marks", 0)
        elements.append(Paragraph(f"<b>Q{i}.</b> {q['question_text']}", q_style))
        elements.append(Paragraph(f"[{marks} marks]", marks_style))
        elements.append(Spacer(1, 3*mm))
        # Answer lines
        elements.append(Paragraph("Answer: _______________________________________________", answer_style))
        elements.append(Spacer(1, 2*mm))
        for _ in range(5):
            elements.append(Paragraph("_______________________________________________________", answer_style))
        elements.append(Spacer(1, 6*mm))

    doc.build(elements)
    return buf.getvalue()


def generate_excel_export(
    summary_headers: List[str],
    summary_rows: List[List],
    detail_headers: List[str],
    detail_rows: List[List],
    exam_title: str = "Exam Results",
) -> bytes:
    """Generate a real Excel (.xlsx) workbook with two sheets.

    Sheet 1 "Summary" holds one row per student (totals + percentage).
    Sheet 2 "Detailed Results" holds the per-question breakdown.

    Returns:
        .xlsx bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    def write_sheet(ws, headers, rows, widths=None):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in rows:
            ws.append(row)
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.border = border
                cell.alignment = wrap
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        if widths:
            for idx, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = w
        else:
            # Auto-size columns from content (capped for readability)
            for idx, col in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col:
                    val = cell.value
                    if val is None:
                        continue
                    length = len(str(val))
                    if length > max_len:
                        max_len = length
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 60)

    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_sheet(ws_summary, summary_headers, summary_rows)

    ws_detail = wb.create_sheet("Detailed Results")
    write_sheet(ws_detail, detail_headers, detail_rows, widths=[10, 22, 26] + [32, 10, 12, 12, 12, 14, 40, 12] * ((len(detail_headers) - 3) // 8) + [12, 14, 12])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_exam_results_pdf(
    exam_title: str,
    subject_name: str,
    students: List[Dict],
    total_marks: float,
) -> bytes:
    """Generate a multi-page PDF report covering every student in an exam.

    Args:
        exam_title: Name of the exam.
        subject_name: Subject the exam belongs to.
        students: List of dicts with keys:
            student_name, student_email, total_score, questions_and_scores
            (list of dicts: question_text, student_answer, score, marks, feedback)
        total_marks: Maximum possible score for the exam.

    Returns:
        PDF bytes.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=20, spaceAfter=4)
    subtitle_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=11, alignment=TA_CENTER, textColor=colors.grey)
    heading_style = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, spaceAfter=4)
    body_style = ParagraphStyle("Body2", parent=styles["Normal"], fontSize=10, leading=14)
    feedback_style = ParagraphStyle("FB", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#4F46E5"), leftIndent=10)
    student_header_style = ParagraphStyle("SH", parent=styles["Heading3"], fontSize=13, textColor=colors.HexColor("#4F46E5"), spaceAfter=4)

    elements = []

    # Cover header
    elements.append(Paragraph("Exam Results Report", title_style))
    elements.append(Paragraph("Smart Exam Answer Checker", subtitle_style))
    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph(f"<b>{exam_title}</b> &nbsp;|&nbsp; {subject_name}", subtitle_style))
    elements.append(Spacer(1, 6*mm))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#4F46E5")))
    elements.append(Spacer(1, 8*mm))

    if not students:
        elements.append(Paragraph("No student results available for this exam.", body_style))
        doc.build(elements)
        return buf.getvalue()

    for i, student in enumerate(students):
        student_name = student.get("student_name", "Unknown")
        student_email = student.get("student_email", "")
        total_score = student.get("total_score", 0)
        percentage = round((total_score / total_marks) * 100, 2) if total_marks > 0 else 0

        # Student header
        elements.append(Paragraph(f"Student {i + 1}: {student_name}", student_header_style))
        info = [
            ["Email:", student_email],
            ["Total Score:", f"{total_score} / {total_marks}"],
            ["Percentage:", f"{percentage}%"],
        ]
        t = Table(info, colWidths=[30*mm, 120*mm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 4*mm))

        # Per-question breakdown
        for q in student.get("questions_and_scores", []):
            elements.append(Paragraph(f"<b>Question:</b> {q.get('question_text', '')[:300]}", body_style))
            elements.append(Paragraph(f"<b>Answer:</b> {q.get('student_answer', '')[:500]}", body_style))
            score_val = q.get("score", 0)
            marks_val = q.get("marks", 0)
            elements.append(Paragraph(f"<b>Score:</b> {score_val} / {marks_val}", body_style))
            fb = q.get("feedback", "")
            if fb:
                elements.append(Paragraph(f"Feedback: {fb}", feedback_style))
            elements.append(Spacer(1, 4*mm))

        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
        elements.append(Spacer(1, 6*mm))

        # Page break between students except after the last one
        if i < len(students) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    return buf.getvalue()
